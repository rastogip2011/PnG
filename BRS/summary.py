from variables import *
import pandas as pd
import openpyxl as op
import os
import shutil

print("\nStep 4:  Updating summary sheet")

vu = dict()
brs = dict()
business = dict()

brslist = pd.read_excel(optimizer_output, sheet_BRS).values
vulist = pd.read_excel(optimizer_output, sheet_VU).values
summarylist = pd.read_excel(filename_summary, sheet_summary).values

for i in vulist:
	try:
		branch = i[1].split(' + ')
		for b in branch:
			br = b.split(' ')[0]
			vu[br] = i[2]
	except:
		a=1

for i in summarylist:
	try:
		branch = i[3].split(' ')[0]
		brs[branch]=0
		business[branch]=i[6]
	except:
		a=1

for i in brslist:
	try:
		branch = i[1].split(' + ')
		s = 0
		for b in branch:
			b=b.split(' ')[0]
			s = s+business[b]
		for b in branch:
			b=b.split(' ')[0]
			brs[b]=brs[b]+i[2]*business[b]/s/1000
	except:
		a=1

try:
	parcellist = pd.read_excel(output_parcel_cost).values
	for i in parcellist[1:]:
		try:
			branch = i[0].split(' ')[0]
			brs[branch]=i[1]/1000
		except:
			a=1
except:
	a=1

# Writing values in summary sheet
 
wbk = op.load_workbook(filename_summary)
wb=wbk[sheet_summary]

n = summarylist.shape[0]
for i in range(2,n+2):
	try:
		branch = wb.cell(i,4).value
		branch = branch.split(' ')[0]

		if branch in brs:
			wb.cell(i,14).value = round(brs[branch], 2)		# BRS
		else:
			wb.cell(i,14).value = 0
		
		if branch in vu:	
			wb.cell(i,23).value = round(vu[branch], 2)		# VU
		else:
			wb.cell(i,23).value = 0							
			
	except:
		a=1

wbk.save(filename_summary)
wbk.close

shutil.copy(filename_summary,summary_copy_path_local)
#shutil.copy(filename_summary,summary_copy_path_drive)

print("    Summary sheet updated.\n")
print("Completed!")
